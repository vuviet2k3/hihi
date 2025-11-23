__author__  = "Vu Van Viet"
__date__    = "11/21/25 - HUST"

"""
Comment: 1/22/25: Done!
Test:
    n GRID INDEPENDENT          : OK!
    PV BUS CONNECTED FOR GRID   : OK!
    n SLACK CONNECTED FOR GIRD  : OK!
    BUS OUT OF SERVICE          : OK!
    LOOP FOR GRID               : OK!
"""

import os, sys, argparse
import openpyxl, math, cmath, time, shutil
import tkinter as tk
from tkinter import messagebox


PATH_PY = os.path.dirname(__file__)
PATH_DATA = os.path.join(PATH_PY, 'data')
PATH_RESULT = os.path.join(PATH_PY, 'result')
PARSER_INPUT = argparse.ArgumentParser(epilog="")
PARSER_INPUT.add_argument('-f', help='(str) Input file.xlsx', default='ieee33.xlsx', type=str)
AGVRS = PARSER_INPUT.parse_known_args()[0]

DATA_xls = os.path.join(PATH_DATA, AGVRS.f)


def warning(title):
    root = tk.Tk()
    root.withdraw()
    messagebox.showwarning("ERROR", title)

def To2S(val):
    return "{:.2f}".format(val)

def To5S(val):
    return "{:.5f}".format(val)

def returnObject(val: str):
    if type(val) == str:
        val = val.strip()
        re_val = val.replace('.', '', 1)
        #
        if re_val.isdigit():
            if '.' in val:
                return float(val)
            else:
                return int(val)
        else:
            return val
    #
    return val

def ReadInput2Setting(work_book=None, sheet_name='SETTING'):
    if work_book is None:
        title = 'WORKBOOK NOT LOADED YET'
        warning(title)
        sys.exit()

    sheet = work_book[sheet_name]
    setting = {}
    #
    for i in range(1, 10000):
        v1 = sheet.cell(i, 1).value
        if v1 is None:
            v2 = sheet.cell(i+1, 1).value
            if v2 is None:
                break
        else:
            if v1.startswith('##'):
                continue
            else:
                for j in range(2, 1000):
                    v2 = sheet.cell(i, j).value
                    if v2:
                        setting.setdefault(str(v1), []).append(returnObject(v2))
                    else:
                        break
    #
    if not setting:
        title = f'NO DATA FOUND SHEET {sheet_name}'
        warning(title)
        sys.exit()
    else:
        return setting

def ReadInput2Sheet(work_book=None, sheet_name=None):
    if work_book is None:
        title = 'WORKBOOK NOT LOADED YET'
        warning(title)
        sys.exit()
    #
    if sheet_name is None:
        title = 'SHEET NOT LOADED YET'
        warning(title)
        sys.exit()

    try:
        sheet = work_book[sheet_name]
    except:
        title = f'SHEET NAME {sheet_name} DOES NOT EXIST.'
        warning(title)
        sys.exit()

    res = {}

    # Raw - bat dau doc data tu hang
    for i in range(1, 10000):
        if i == 9999:
            title = f'NO KEY DATA "##" FOUND SHEET NAME: {sheet_name}'
            warning(title)
            sys.exit()
        #
        if sheet.cell(i, 1).value.startswith('##'):
            Raw = i
            break

    # read data
    res = {}
    for i in range(Raw+1, 10000):
        if i == 9999:
            title = f'NO DATA FOUND IN SHEET NAME: {sheet_name}'
            warning(title)
            sys.exit()
        #
        res1 = {}
        v1 = sheet.cell(i+1, 1).value
        if v1:
            for j in range(2, 10000):
                v2 = sheet.cell(Raw+1, j).value
                if v2:
                    v3 = sheet.cell(i+1, j).value
                    res1[v2] = returnObject(v3)
                else:
                    break
            res[returnObject(v1)] = res1
        else:
            break
    #
    return res

# def WriteOut2Sheet(sheet_book=None, data=None):


class DATA:
    global DATA_xls

    def __init__(self, input=DATA_xls):
        wb = openpyxl.load_workbook(input, data_only=True)
        self.Asetting = ReadInput2Setting(work_book=wb, sheet_name='SETTING')
        self.Abus = ReadInput2Sheet(work_book=wb, sheet_name='BUS')
        self.Aline = ReadInput2Sheet(work_book=wb, sheet_name='LINE')
        self.Asource = ReadInput2Sheet(work_book=wb, sheet_name='SOURCE')
        self.Atrf2 = ReadInput2Sheet(work_book=wb, sheet_name='TRF2')
        self.Atrf3 = ReadInput2Sheet(work_book=wb, sheet_name='TRF3')
        self.Ashunt = ReadInput2Sheet(work_book=wb, sheet_name='SHUNT')
        self.data()

    def data(self):
        """
        self.AlgoPF          : [PF, nmax, eps]
        self.Sbase
        self.Bus             : Bus ID      
        self.BrnC0           : {brnID:[frombus, tobus]}
        self.BrnC1           : {bus: [brnID]}
        trf2 : ID = ID + len(lineID)
        """
        # SETTING
        unit = self.Asetting['GE_PowerUnit'][0]
        if unit.upper() not in {'MVA','KVA'}:
            title = f'POWER UNIT not in MVA, KVA'
            warning(title)
            sys.exit()
        #
        if unit.upper() == 'MVA':
            self.sbase = self.Asetting['GE_Sbase'][0] * 1e6
        else:
            self.sbase = self.Asetting['GE_Sbase'][0] * 1e3
        self.AlgoPF = self.Asetting['PF']

        # BUS
        self.busAll = {}
        memo = 1
        for key, v in self.Abus.items():
            if v['FLAG'] == 1:
                unit = v['MEMO']
                if unit:
                    if unit.upper() not in {'MVA', 'KVA'}:
                        title = f'MEMO SHEET BUS not in MVA, KVA'
                        warning(title)
                        sys.exit()
                    #
                    if unit.upper() == 'MVA':
                        memo = memo * 1e6
                    elif unit.upper() == 'KVA':
                        memo = memo * 1e3
                #
                v1 = v['PLOAD'] if v['PLOAD'] else 0
                v2 = v['QLOAD'] if v['QLOAD'] else 0
                v1pu = v1 * memo / self.sbase
                v2pu = v2 * memo / self.sbase
                self.busAll[key] = [v1pu, v2pu, v['kV']]

        # SOURCE
        self.slackAll = {}
        self.pvAll = {}
        memo = 1
        for key, v in self.Asource.items():
            if v['FLAG'] == 1:
                unit = v['MEMO']
                if unit:
                    if unit.upper() not in {'MVA', 'KVA'}:
                        title = f'MEMO SHEET SOURCE not in MVA, KVA'
                        warning(title)
                        sys.exit()
                    #
                    if unit.upper() == 'MVA':
                        memo = memo * 1e6
                    else:
                        memo = memo * 1e3
                #
                if v['CODE'] == 3:
                    v1 = v['BUS_ID']
                    v2 = v['vGen [pu]']
                    v3 = v['aGen [deg]'] *math.pi / 180
                    self.slackAll[v1] = [v2, v3]
                elif v['CODE'] == 2:
                    v1 = v['BUS_ID']
                    v2 = v['vGen [pu]']
                    v3 = v['Pgen'] * memo / self.sbase
                    self.pvAll[v1] = [v2, v3]

        # LINE
        self.brnC0 = {}
        self.brnC1 = {}
        self.lineAll = {}
        for k, v in self.Aline.items():
            if v['FLAG'] == 1:
                b1, b2 = v['BUS_ID1'], v['BUS_ID2']
                self.brnC0[k] = [b1, b2]
                self.brnC1.setdefault(b1, []).append(k)
                self.brnC1.setdefault(b2, []).append(k)
                #
                l = v['LENGTH [km]']
                kv = v['kV']
                r = v['R [Ohm/km]'] * l * self.sbase / (kv * 1e3)**2
                x = v['X [Ohm/km]'] * l * self.sbase / (kv * 1e3)**2
                b = v['B [microS/km]'] * l * (kv * 1e3)**2 / self.sbase / 1e6
                self.lineAll[k] = [complex(r, x), complex(0, b)]

        # TRF2
        # Neu co trf2 thi trf2ID se trung voi lineID
        # ->  New trf2ID = trf2ID + len(lineID)
        kline = len(self.lineAll.keys())
        self.x2All = {}
        memo1, memo2 = 1, 1
        for k, v in self.Atrf2.items():
            if v['FLAG'] == 1:
                b1, b2 = v['BUS_ID1'], v['BUS_ID2']
                self.brnC0[k + kline] = [b1, b2]
                self.brnC1.setdefault(b1, []).append(k + kline)
                self.brnC1.setdefault(b2, []).append(k + kline)
                #
                unit = v['MEMO']
                if unit:
                    p1, p2 = unit.split(',')
                    p1, p2 = p1.strip(), p2.strip()
                    if p1.upper() not in {'MVA', 'KVA'}:
                        title = 'MEMO S(VA) SHEET TRF2 not in MVA, KVA'
                        warning(title)
                        sys.exit()
                    elif p1.upper() == 'MVA':
                        memo1 = memo1 * 1e6
                    elif p1.upper() == 'KVA':
                        memo1 = memo1 * 1e3
                    #
                    if p2.upper() not in {'MW', 'KW'}:
                        title = 'MEMO P(W) SHEET TRF2 not in MW, KW'
                        warning(title)
                        sys.exit()
                    elif p2.upper() == 'MW':
                        memo2 = memo2 * 1e6
                    elif p2.upper() == 'KW':
                        memo2 = memo2 * 1e3

                sn = v['Sn'] * memo1
                r = v['pk'] * memo2 * self.sbase / sn**2
                x = v['uk [%]'] * self.sbase / sn / 1e2
                g = v['P0'] * memo2 / self.sbase
                b = v['i0 [%]'] * sn / self.sbase / 1e2
                #
                kv1, kv2 = v['kV1'], v['kV2']
                if kv1 > kv2:
                    self.x2All[k + kline] = [b1, complex(r, x), complex(g, -b)]
                else:
                    self.x2All[k + kline] = [b2, complex(r, x), complex(g, -b)]

        # TRF3

        # SHUNT
        self.shuntAll = {}
        memo = 1
        for k, v in self.Ashunt.items():
            if v['FLAG'] == 1:
                v1 = v['BUS_ID']
                unit = v['MEMO']
                if unit:
                    if unit.upper() not in {'MVAR', 'KVAR'}:
                        title = 'MEMO VAR SHEET SHUNT not in MVAR, KVAR'
                        warning(title)
                        sys.exit()
                    elif unit.upper() == 'MVAR':
                        memo = memo * 1e6
                    elif unit.upper() == 'KVAR':
                        memo = memo * 1e3

                g = v['deltaP'] * memo / self.sbase
                b = v['Qshunt'] * memo / self.sbase
        
class PSM:
    def __init__(self,
                 abus=None,
                 aslack=None,
                 apv=None,
                 brnC0=None,
                 brnC1=None,
                 aline=None,
                 atrf2=None,
                 ashunt=None,
                 nMax=200,
                 Eps=1e-5):
        """
        Run Power Summation Method - PSM
        YEU CAU : thong so he don vi TUONG DOI - p.u
        abus    : {busID: [pLoad, qLoad]}
        aslack  : {busID: [vGen, aGen]}
        apv     : {busID: [vGen, pGen]}
        brnC0   : {brnID: [frombus, tobus]}
        brnC1   : {bus: [brnID]}
        trf2    : ID> 100000
        aline   : {lineID: [rLine + 1j*xLine, bLine]}
        atrf2   : {x2ID: [rX2 + 1j*xX2, gX2 - 1jbX2]}
        ashunt  : {shuntID: [gShunt + 1j*bShunt]}

        :return -> ubus, abus
        """
        self.abus = abus
        self.aslack = aslack
        self.apv = apv
        self.brnC0 = brnC0
        self.brnC1 = brnC1
        #
        self.aline = aline
        self.atrf2 = atrf2
        self.ashunt = ashunt
        self.nMax = nMax
        self.Eps = Eps

    # check loop
    def check_loop(self, slack):
        """
        CHECK LOOP: THUAT TOAN DFS ket hop de quy
        return True -> LOOP
        return False -> NO LOOP
        """
        visited = set()
        visitID = set()
        #
        def dfs(node, parent):
            visited.add(node)
            for line in self.brnC1[node]:
                if line in visitID:
                    continue
                visitID.add(line)
                bus1, bus2 = self.brnC0[line]
                bus = bus1 if bus1 != node else bus2
                if bus in visited and bus != parent:
                    return True
                if dfs(bus, node):
                    return True
            return False

        return dfs(slack, None)

    # mapping
    def mapping(self, slack):
        mapping = list()
        visited = set()

        def dfs(bus):
            mapping.append(bus)
            for line in self.brnC1[bus]:
                if line not in visited:
                    visited.add(line)
                    v1, v2 = self.brnC0[line]
                    v = v1 if v1 != bus else v2
                    dfs(v)

        dfs(slack)
        return mapping
    
    # check source
    def check_source(self, slack, pv, mapping):
        for k in pv.keys():
            if k in mapping:
                return f'PV BUS{k} CONNECTED TO THE MAIN GRID. PLEASE USE A DIFFERENT METHOD!!!'
        #
        for k in self.aslack.keys():
            if k == slack:
                continue
            if k in mapping:
                return f'SLACK BUS{k} AND SLACK BUS{slack} CONNECTED TO THE GRID. PLEASE CHECK THE NETWORK SERVICE!!!'
        return None
    
    # check bus service
    def check_service(self, bus, mapping):
        for k in bus.keys():
            if k not in mapping:
                return k 
        return None


    def power_shunt(self, ubus, mapping):
        """
        Tinh toan POWER BUS co xet thanh phan SHUNT
        ubus: {busID: volt}

        :return -> sbus
        """
        sbus = {}
        for k in mapping:
            v = self.abus[k]
            sbus[k] = complex(v[0], v[1])
            s = complex(0, 0)
            #
            for brn in self.brnC1[k]:
                if brn in self.aline:
                    line = self.aline[brn]
                    s += line[1] / 2
                #
                if brn in self.atrf2:
                    x2 = self.atrf2[brn]
                    if k == x2[0]:
                        s += x2[2]
            #
            if k in self.ashunt:
                s += self.ashunt[k]
            sbus[k] += s.conjugate() * abs(ubus[k]) ** 2

        return sbus

    def backward_sweep(self, ubus, sbus, slack, re_mapping):
        """
        Tinh dong cong suat nguoc - backward sweep

        :return -> sbrn1
        """
        sbrn1 = {}
        sbrn2 = {}
        visited = set()

        def dfs(lineID, visited):
            fromline = int()
            toline = list()
            for line in lineID:
                if line not in visited:
                    fromline = line
                else:
                    toline.append(line)
            return fromline, toline

        for bus in re_mapping:
            if bus == slack:
                break
            lineID = self.brnC1[bus]
            fromline, toline = dfs(lineID, visited)
            sbrn2[fromline] = sbus[bus]
            for line in toline:
                sbrn2[fromline] += sbrn1[line]
            sbrn1[fromline] = sbrn2[fromline] +abs(sbrn2[fromline])**2 / abs(ubus[bus])**2 * self.aline[fromline][0]
            visited.add(fromline)

        return sbrn1


    def forward_sweep(self, ubus, sbrn1, mapping):
        """
        Tinh dien ap thuan - forward sweep

        :return -> ubus
        """
        ubusN=  ubus.copy()
        visited = set()
        #
        for bus in mapping:
            lineID = self.brnC1[bus]
            for line in lineID:
                if line not in visited:
                    v = self.brnC0[line]
                    b1, b2 = v[0], v[1]
                    b = b1 if b1 != bus else b2
                    ubusN[b] = ubusN[bus] - sbrn1[line].conjugate() * self.aline[line][0] / ubusN[bus].conjugate()
                    visited.add(line)

        return ubusN

    def epsilon(self, ubus, ubusN):
        for k, v in ubus.items():
            real1, imag1 = ubus[k].real, ubus[k].imag
            real2, imag2 = ubusN[k].real, ubusN[k].imag
            eps = max(abs(real1 - real2), abs(imag1 - imag2))
            if eps > self.Eps:
                return False
        return True


    def solve1slack(self, slack, param, mapping, re_mapping):
        # initialize bus
        ubus = {}
        for k in mapping:
            ubus[k] = complex(1, 0)
        ubus[slack] = complex(param[0], param[1])

        # solve
        for iter in range(1, self.nMax+1):
            if iter == self.nMax:
                title = 'KHONG HOI TU!!!'
                warning(title)
                sys.exit()

            sbus = self.power_shunt(ubus, mapping)
            sbrn1 = self.backward_sweep(ubus, sbus, slack, re_mapping)
            ubusN = self.forward_sweep(ubus, sbrn1, mapping)

            if self.epsilon(ubus, ubusN):
                for k in mapping:
                    v = self.abus[k]
                    kv = v[2]
                    ubusN[k] = [abs(ubusN[k])*kv, abs(ubusN[k]), cmath.phase(ubusN[k])] 
                return ubusN

            else:
                ubus = ubusN.copy()
                iter += 1
        
    def solve(self):
        t_start = time.time()
        Allbus = set(self.abus.keys())
        serve_bus = set()
        ubus = {}
        # for slack in self.aslack.keys():
        for k, v in self.aslack.items():
            slack = k
            param = v

            # check_loop
            if self.check_loop(slack=slack):
                title = 'LOOP DETECTED IN THE NETWORK. PLEASE USE A DIFFERENT METHOD!!!'
                warning(title)
                sys.exit()

            # mapping for grid
            mapping = list(self.mapping(slack=slack))
            re_mapping = mapping[::-1]

            # check source
            if self.apv is not None:
                k = self.check_source(slack, self.apv, mapping)
                if k is not None:
                    title = k
                    warning(title)
                    sys.exit()

            # check service
            # k = self.check_service(self.abus, mapping)
            # if k is not None:
            #     title = f'BUS{k} NOT CONNECTED TO THE MAIN GRID. PLEASE CHECK THE NETWORK SERVICE!!!'
            #     warning(title)
            #     sys.exit()

            # solve for 1 slack
            ubus.update(self.solve1slack(slack, param, mapping, re_mapping))

            serve_bus |= set(mapping)

        # check service
        if serve_bus != Allbus:
            diff_bus = Allbus - serve_bus
            title = f'BUS{diff_bus.pop()} NOT CONNECTED TO THE GRID. PLEASE CHECK THE NETWORK SERVICE!!!'
            warning(title)
            sys.exit()
        
        return ubus, time.time() - t_start
            


class PSSe:
    def __init__(self,
                sbase=100 * 1e6,
                abus=None,
                aslack=None,
                apv = None,
                brnC0=None,
                brnC1=None,
                aline=None,
                atrf2=None,
                ashunt=None):

        self.sbase = sbase
        self.abus = abus
        self.aslack = aslack
        self.apv = apv
        self.brnC0 = brnC0
        self.brnC1 = brnC1
        #
        self.aline = aline
        self.atrf2 = atrf2
        self.ashunt = ashunt

        PY_VER = str(sys.version_info[0]) + '.' + str(sys.version_info[1])
        if PY_VER != '3.11':
            title = f'PLEASE USE PYTHON 3.11.x VERSION TO PSS/E 36 XPLORE!!!\n CURRENT VERSION IS {PY_VER}'
            warning(title)
            sys.exit()



    def solve(self):
        import psse36
        import psspy

        #
        _i = psspy.getdefaultint()
        _f = psspy.getdefaultreal()
        _s = psspy.getdefaultchar()

        # initialize PSS/E
        ierr = psspy.psseinit()
        fHz = 50  # Hz
        ierr = psspy.newcase_2([0,1], self.sbase/1e6, fHz, "hihi", "haha")

        # add buses and load
        for k, v in self.abus.items():
            p = v[0] * self.sbase / 1e6  # MW
            q = v[1] * self.sbase / 1e6  # MVAR
            kv = v[2] # kV
            #
            code = 1
            if k in self.aslack:
                code = 3
            elif k in self.apv:
                code = 2
            ierr = psspy.bus_data_4(k, 0, [code, _i, _i, _i], [kv, _f, _f, _f, _f, _f, _f], f"BUS{k}")
            #
            ierr = psspy.load_data_7(k, "1", [_i]*7, [p, q, _f, _f, _f, _f, _f, _f], "", "")

        # add source
        for k, v in self.aslack.items():
            # ierr = psspy.machine_data_5(k, "1", )
            ierr = psspy.plant_data_4(k, 0, [_i]*2, [v[0], _f])
        #
        if self.apv is not None:
            for k, v in self.apv.items():
                ierr = psspy.plant_data_4(k, 0, [_i]*2, [v[0], _f])
                ierr = psspy.machine_data_5(k, "1", [_i]*5, [v[1] * self.sbase/1e6] + [_f]*16, ["", ""])
        
        # add line
        for k, v in self.aline.items():
            fbus, tbus = self.brnC0[k]
            r = v[0].real
            x = v[0].imag
            g = v[1].real
            b = v[1].imag
            ierr = psspy.branch_data_4(fbus, tbus, "1", [_i]*7, [r, x, g, b] + [_f]*8, [_f]*12, f"BUS{fbus}-BUS{tbus}")

        # add trf2
        
        # add shunt

        # add trf3

        # slove
        ierr = psspy.fnsl([0,0,0,1,1,0,99,0])

        #get data NUMBER, NAME, PU, kV, ANGLE
        ierr, num_bus = psspy.abusint(-1, _i, 'NUMBER')
        ierr, name_bus = psspy.abuschar(-1, _i, 'NAME')
        ierr, pu_bus = psspy.abusreal(-1, _i, 'PU')
        ierr, kv_bus = psspy.abusreal(-1, _i, 'KV')
        ierr, angle_bus = psspy.abusreal(-1, _i, 'ANGLE')
        num_bus = num_bus[0]
        name_bus = name_bus[0]
        pu_bus = pu_bus[0]
        kv_bus = kv_bus[0]
        angle_bus = angle_bus[0]

        # save to sav/raw
        filename = os.path.splitext(os.path.basename(DATA_xls))[0]
        file_sav = os.path.join(PATH_RESULT, f"{filename}.sav")
        file_raw = os.path.join(PATH_RESULT, f"{filename}.raw")
        ierr = psspy.save(file_sav)
        ierr = psspy.rawd(_i, _i, [_i]*6, 0, file_raw)

        return num_bus, name_bus, kv_bus, pu_bus, angle_bus



def run():
    pf = DATA()
    nMax = pf.AlgoPF[1]
    Eps = pf.AlgoPF[2]
    psm = PSM(
                abus = pf.busAll,
                aslack = pf.slackAll,
                apv = pf.pvAll,
                brnC0 = pf.brnC0,
                brnC1 = pf.brnC1,
                aline = pf.lineAll,
                atrf2 = pf.x2All,
                ashunt = pf.shuntAll,
                nMax = nMax,
                Eps = Eps
            )
    bus, time = psm.solve()
    #
    psse = PSSe(
                sbase = pf.sbase,
                abus = pf.busAll,
                aslack = pf.slackAll,
                apv = pf.pvAll,
                brnC0 = pf.brnC0,
                brnC1 = pf.brnC1,
                aline = pf.lineAll,
                atrf2 = pf.x2All,
                ashunt = pf.shuntAll
            )
    num_bus, name_bus, kv_bus, pu_bus, angle_bus = psse.solve()

    ###### write output to sheet BUS out.xlsx
    filename = os.path.splitext(os.path.basename(DATA_xls))[0]
    outfile = os.path.join(PATH_RESULT, f"OUT_{filename}.xlsx")
    examfile = os.path.join(PATH_RESULT, "example", "example.xlsx")
    try:
        shutil.copyfile(examfile, outfile)
    except:
        title = f'PLEASE CLOSE FILE: {outfile}'
        warning(title)
        sys.exit()

    #
    wb = openpyxl.load_workbook(outfile)
    ws = wb['BUS']

    # write time
    ws.cell(2, 2).value = time

    # write methor & pss/e
    row = 7
    col = 1
    for k, v in bus.items():
        i = num_bus.index(k)
        #
        ws.cell(row, col).value = k
        ws.cell(row, col+1).value = name_bus[i]
        #
        ws.cell(row, col+2).value = To2S(v[0])
        ws.cell(row, col+3).value = To5S(v[1])
        ws.cell(row, col+4).value = To5S(v[2]*180/math.pi)
        #   
        ws.cell(row, col+5).value = To2S(kv_bus[i])
        ws.cell(row, col+6).value = To5S(pu_bus[i])
        ws.cell(row, col+7).value = To5S(angle_bus[i]*180/math.pi)
        #
        ws.cell(row, col+8).value = To2S(abs(v[0] - kv_bus[i])/kv_bus[i]*100)
        ws.cell(row, col+9).value = To2S(abs(v[1] - pu_bus[i])/pu_bus[i]*100)
        ws.cell(row, col+10).value = To2S(abs(v[2] - angle_bus[i])/abs(angle_bus[i])*100) if angle_bus[i] != 0 else To2S(0)

        row += 1

    wb.save(outfile)
    os.startfile(outfile)


if __name__ == "__main__":
    run()

